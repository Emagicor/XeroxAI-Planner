"""
infrastructure/exporters/csv_exporter.py + xlsx_exporter.py

Moved from routes/export.py into infrastructure where exporters belong.
Routes now call these functions and return the bytes as a Response.
"""
from __future__ import annotations

import csv
import io

from domain.entities.job import AnalyzeJob
from infrastructure.exporters.formatter import build_summary, build_table_rows


# ── Shared column spec ────────────────────────────────────────────────────────

def _headers(unit_label: str) -> list[str]:
    return [
        "Page/Floor",
        "Area Name",
        "Dimensions",
        f"Computed Area ({unit_label})",
        "Method",
        "Confidence %",
        "Assumed?",
        "Notes",
    ]


def _row_values(r: dict) -> list:
    return [
        r["floor"],
        r["area_name"],
        r["dimensions"],
        r["computed_area"] if r["computed_area"] is not None else "",
        r["method"],
        r["confidence_pct"],
        "Yes" if r["assumed"] else "No",
        r["notes"],
    ]


# ── CSV ───────────────────────────────────────────────────────────────────────

def export_csv(job: AnalyzeJob, unit: str = "sqft") -> bytes:
    rows    = build_table_rows(job, unit)
    summary = build_summary(job, unit)

    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(_headers(summary["unit"]))

    for r in rows:
        w.writerow(_row_values(r))

    w.writerow([])
    w.writerow(["Grand Total", "", "", summary["grand_total"], summary["unit"], "", "", ""])
    w.writerow(["Eligible pages",   summary["eligible_pages"]])
    w.writerow(["Ineligible pages", summary["ineligible_pages"]])

    return buf.getvalue().encode("utf-8-sig")   # BOM for Excel compat


# ── XLSX ──────────────────────────────────────────────────────────────────────

def export_xlsx(job: AnalyzeJob, unit: str = "sqft") -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows    = build_table_rows(job, unit)
    summary = build_summary(job, unit)
    headers = _headers(summary["unit"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Floor Plan Areas"

    # Styles
    hdr_fill   = PatternFill("solid", fgColor="1E3A5F")
    hdr_font   = Font(bold=True, color="FFFFFF")
    assumed_fill    = PatternFill("solid", fgColor="FFF3CD")
    ineligible_fill = PatternFill("solid", fgColor="F8D7DA")
    bold = Font(bold=True)

    ws.append(headers)
    for ci, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci)
        c.font  = hdr_font
        c.fill  = hdr_fill
        c.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append(_row_values(r))
        ri = ws.max_row
        fill = ineligible_fill if not r.get("eligible", True) else (
            assumed_fill if r.get("assumed") else None
        )
        if fill:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).fill = fill

    ws.append([])
    ws.append(["Grand Total", "", "", summary["grand_total"], summary["unit"], "", "", ""])
    for ci in range(1, 5):
        ws.cell(row=ws.max_row, column=ci).font = bold

    ws.append(["Eligible pages",   summary["eligible_pages"]])
    ws.append(["Ineligible pages", summary["ineligible_pages"]])

    # Auto-fit columns
    for ci in range(1, len(headers) + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            (len(str(ws.cell(row=ri, column=ci).value or ""))
             for ri in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()