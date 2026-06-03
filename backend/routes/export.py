"""
routes/export.py
----------------
POST /export/csv   → returns a CSV file
POST /export/xlsx  → returns an XLSX file

Body (JSON):
  {
    "analysis": <full /analyze response>,
    "unit": "sqft" | "sqm" | "sq-in" | "sq-cm"   (default "sqft")
  }
"""

from __future__ import annotations

import csv
import io

from fastapi import APIRouter
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from typing import Any

from output.formatter import build_summary, build_table_rows

export_router = APIRouter()

COLUMNS = [
    "page",
    "floor",
    "area_name",
    "dimensions",
    "computed_area",
    "unit",
    "method",
    "confidence_pct",
    "assumed",
    "notes",
]

COLUMN_HEADERS = [
    "Page/Floor",
    "Area Name",
    "Dimensions",
    "Computed Area",
    "Unit",
    "Method",
    "Confidence %",
    "Assumed?",
    "Notes",
]


class ExportRequest(BaseModel):
    analysis: dict[str, Any]
    unit: str = "sqft"


# ── CSV ───────────────────────────────────────────────────────────────────────

@export_router.post("/export/csv")
def export_csv(req: ExportRequest):
    rows = build_table_rows(req.analysis, req.unit)
    summary = build_summary(req.analysis, req.unit)

    buf = io.StringIO()
    writer = csv.writer(buf)

    # Header
    writer.writerow(["Page/Floor", "Area Name", "Dimensions",
                     f"Computed Area ({summary['unit']})", "Method",
                     "Confidence %", "Assumed?", "Notes"])

    for r in rows:
        writer.writerow([
            r["floor"],
            r["area_name"],
            r["dimensions"],
            r["computed_area"] if r["computed_area"] is not None else "",
            r["method"],
            r["confidence_pct"],
            "Yes" if r["assumed"] else "No",
            r["notes"],
        ])

    # Summary footer
    writer.writerow([])
    writer.writerow(["Grand Total", "", "", summary["grand_total"],
                     summary["unit"], "", "", ""])
    writer.writerow(["Eligible pages", summary["eligible_pages"], "", "",
                     "", "", "", ""])
    writer.writerow(["Ineligible pages", summary["ineligible_pages"], "", "",
                     "", "", "", ""])

    content = buf.getvalue().encode("utf-8-sig")  # BOM for Excel compat
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=floor_plan_areas.csv"},
    )


# ── XLSX ──────────────────────────────────────────────────────────────────────

@export_router.post("/export/xlsx")
def export_xlsx(req: ExportRequest):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl",
        ) from exc

    rows = build_table_rows(req.analysis, req.unit)
    summary = build_summary(req.analysis, req.unit)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Floor Plan Areas"

    # ── styles ──
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    assumed_fill = PatternFill("solid", fgColor="FFF3CD")
    ineligible_fill = PatternFill("solid", fgColor="F8D7DA")
    total_font = Font(bold=True)

    headers = [
        "Page/Floor", "Area Name", "Dimensions",
        f"Computed Area ({summary['unit']})",
        "Method", "Confidence %", "Assumed?", "Notes",
    ]

    # Write header row
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for r in rows:
        row_values = [
            r["floor"],
            r["area_name"],
            r["dimensions"],
            r["computed_area"],
            r["method"],
            r["confidence_pct"],
            "Yes" if r["assumed"] else "No",
            r["notes"],
        ]
        ws.append(row_values)
        current_row = ws.max_row

        # Colour ineligible rows
        if not r.get("eligible", True):
            for col in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=col).fill = ineligible_fill
        elif r.get("assumed"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=col).fill = assumed_fill

    # Summary footer
    ws.append([])
    total_row = ws.max_row + 1
    ws.append(["Grand Total", "", "",
                summary["grand_total"], summary["unit"], "", "", ""])
    for col in range(1, 5):
        ws.cell(row=ws.max_row, column=col).font = total_font

    ws.append(["Eligible pages", summary["eligible_pages"]])
    ws.append(["Ineligible pages", summary["ineligible_pages"]])

    # Auto-fit columns
    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(row=r_idx, column=col_idx).value or ""))
             for r_idx in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=floor_plan_areas.xlsx"},
    )